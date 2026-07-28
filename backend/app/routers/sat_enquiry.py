from __future__ import annotations

import asyncio
import io
import json
import os
import re
import secrets
from datetime import datetime
from typing import Any, NamedTuple
from uuid import uuid4

import httpx
import openpyxl
import redis
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel

from ..credentials import get as get_cred
from ..rate_limits import rate_limit
from ..security import verify_user_identity
from ..services.error_reporting import report
from ..middleware import SAT_XLSX_MAX_BYTES

router = APIRouter(prefix="/api/sat", tags=["sat"])

_DIVERZA_BASE = "https://servicios.diverza.com/api/v2/documents"
_PRIORITY_FIELDS = {"estatus_cancelacion", "estado", "es_cancelable"}

# El UUID entra crudo a una URL de un tercero que se llama AUTENTICADA con el
# credential_id/credential_token del emisor. No basta con confiar en que httpx
# lo escape: httpx normaliza "../" según RFC 3986, así que un uuid de
# "../../../admin" convierte
#   https://servicios.diverza.com/api/v2/documents/{uuid}/sat_cfdi_enquiry
# en
#   https://servicios.diverza.com/admin/sat_cfdi_enquiry
# (comprobado con httpx.URL el 2026-07-26). Se valida la forma, que es la
# única defensa que no depende del comportamiento de la librería.
_UUID_RE = re.compile(
    r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-"
    r"[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$"
)


def _is_uuid(value: str) -> bool:
    return bool(value) and bool(_UUID_RE.match(value))


def _require_uuid(value: str) -> str:
    if not _is_uuid(value):
        raise HTTPException(status_code=400, detail="UUID de CFDI inválido")
    return value

# Almacén de resultados de consulta por lote.
#
# Antes esto era `_job_results: dict[str, bytes]`, un dict a nivel de módulo, y
# tenía DOS problemas distintos:
#
#   1. Seguridad (#3, CRITICAL). El job_id viajaba en el evento SSE `done` y
#      GET /enquiry/batch/{job_id}/result hacía pop() sin comprobar de quién era.
#      Cualquiera que viera un job_id se llevaba el Excel de otra sesión, con
#      los UUID y RFC de las facturas de otro contribuyente.
#
#   2. Corrección, que la auditoría no menciona. El servicio corre con
#      --max-instances=10: el dict vive en UNA instancia, así que la descarga
#      sólo funcionaba si la petición caía por casualidad en la misma que
#      generó el Excel. Y evictaba con 5 entradas (#18), tirando resultados
#      que el usuario todavía no había bajado.
#
# Ahora el Excel vive en Redis con TTL, y para bajarlo hace falta un token
# aleatorio de 32 bytes que NO es el job_id: el job_id sigue viajando por el
# SSE para el progreso, pero ya no sirve para descargar nada.
_RESULT_TTL_SECONDS = 900  # 15 min: alcanza para bajarlo, se limpia solo

# Cliente propio en vez de reusar el de batch.py: aquel usa
# decode_responses=True y corrompería los bytes binarios del .xlsx.
_redis_client = redis.Redis(
    host=os.getenv("REDIS_HOST", "localhost"),
    port=int(os.getenv("REDIS_PORT", "6379")),
    password=os.getenv("REDIS_PASSWORD", None),
    ssl=True,
    ssl_cert_reqs="required",
    max_connections=10,
    decode_responses=False,
)


def _result_key(job_id: str) -> str:
    return f"sat_enquiry:result:{job_id}"


def _token_key(token: str) -> str:
    return f"sat_enquiry:token:{token}"


# ---------------------------------------------------------------------------
# Pydantic models
# ---------------------------------------------------------------------------


class EnquiryRequest(BaseModel):
    uuid: str
    rfc_emisor: str
    rfc_receptor: str
    total_cfdi: str
    motive: str = "01"


class EnquiryResult(BaseModel):
    uuid: str
    estado: str
    es_cancelable: str
    estatus_cancelacion: str
    error: str | None = None


# ---------------------------------------------------------------------------
# Diverza response parsing
# ---------------------------------------------------------------------------


def _extract_json_objects(text: str) -> list[str]:
    """Extract JSON object strings from arbitrary text (handles malformed responses)."""
    objs: list[str] = []
    start: int | None = None
    depth = 0
    for i, ch in enumerate(text):
        if ch == "{":
            if depth == 0:
                start = i
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0 and start is not None:
                objs.append(text[start : i + 1])
                start = None
    return objs


def _choose_best_json(text: str) -> dict[str, Any] | None:
    best: dict[str, Any] | None = None
    best_score = -1
    for candidate in _extract_json_objects(text):
        try:
            parsed: dict[str, Any] = json.loads(candidate)
            score = sum(k in parsed for k in _PRIORITY_FIELDS)
            if score > best_score:
                best, best_score = parsed, score
        except json.JSONDecodeError:
            pass
    return best


def _parse_diverza_response(text: str) -> dict[str, str | None]:
    parsed = _choose_best_json(text)
    if not parsed:
        return {
            "estado": "",
            "es_cancelable": "",
            "estatus_cancelacion": "",
            "error": "JSON inválido en respuesta de Diverza",
        }

    estado = (parsed.get("estado") or "").strip()
    es_cancelable = (parsed.get("es_cancelable") or "").strip()
    estatus_cancelacion = (parsed.get("estatus_cancelacion") or "").strip()

    # Domain rule from reference implementation
    if (
        estado.lower() == "vigente"
        and es_cancelable.lower() == "no cancelable"
        and not estatus_cancelacion
    ):
        estatus_cancelacion = "No cancelable estatus"

    return {
        "estado": estado,
        "es_cancelable": es_cancelable,
        "estatus_cancelacion": estatus_cancelacion,
        "error": None,
    }


# ---------------------------------------------------------------------------
# Diverza HTTP call
# ---------------------------------------------------------------------------


def _build_payload(
    uuid: str,
    rfc_emisor: str,
    rfc_receptor: str,
    total_cfdi: str,
    motive: str,
    cred: dict[str, str],
) -> dict[str, Any]:
    return {
        "credentials": {
            "id": cred.get("credential_id", ""),
            "token": cred.get("credential_token", ""),
        },
        "issuer": {"rfc": rfc_emisor},
        "document": {
            "certificate-number": cred.get("certificate_number", ""),
            "rfc_receptor": rfc_receptor,
            "total_cfdi": total_cfdi,
            "motive": str(motive).zfill(2),
            "replacement-folio": "",
        },
    }


async def _call_diverza(
    client: httpx.AsyncClient,
    uuid: str,
    payload: dict[str, Any],
    max_retries: int = 3,
) -> str:
    url = f"{_DIVERZA_BASE}/{_require_uuid(uuid)}/sat_cfdi_enquiry"
    last_exc: Exception | None = None

    for attempt in range(1, max_retries + 1):
        try:
            resp = await client.put(url, json=payload, timeout=30.0)
            if 500 <= resp.status_code < 600 and attempt < max_retries:
                await asyncio.sleep(2**attempt)
                continue
            resp.raise_for_status()
            return resp.text
        except httpx.HTTPError as exc:
            last_exc = exc
            if attempt < max_retries:
                await asyncio.sleep(2**attempt)

    raise last_exc or RuntimeError("Max retries exceeded")


async def _enquiry_indexed(
    client: httpx.AsyncClient,
    idx: int,
    uuid: str,
    rfc_emisor: str,
    rfc_receptor: str,
    total_cfdi: str,
    motive: str,
    tenant_id: str,
) -> tuple[int, dict[str, Any]]:
    cred = get_cred(rfc_emisor.upper(), tenant_id)
    if not cred:
        return idx, {
            "uuid": uuid,
            "estado": "",
            "es_cancelable": "",
            "estatus_cancelacion": "",
            "error": f"RFC emisor no configurado: {rfc_emisor}",
        }

    payload = _build_payload(uuid, rfc_emisor, rfc_receptor, total_cfdi, motive, cred)

    try:
        text = await _call_diverza(client, uuid, payload)
    except Exception as exc:
        report(exc, contexto="consulta_sat_lote")
        return idx, {
            "uuid": uuid,
            "estado": "",
            "es_cancelable": "",
            "estatus_cancelacion": "",
            "error": "Error al consultar el SAT",
        }

    result = _parse_diverza_response(text)
    return idx, {"uuid": uuid, **result}


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------


class ParsedInput(NamedTuple):
    """Filas utilizables + cuántas se descartaron por traer un UUID malformado.

    `descartadas` NO cuenta las filas con UUID vacío: en un Excel real las
    últimas filas suelen venir en blanco y reportarlas sería ruido. Cuenta sólo
    las que traen algo que no es un UUID -- ésas sí son un dato del usuario que
    no se va a consultar, y merece decírselo en vez de tragarlo.
    """

    rows: list[dict[str, str]]
    descartadas: int


def _parse_excel_input(content: bytes) -> ParsedInput:
    # read_only=True: el modelo eager de openpyxl expande 10 MB de XLSX
    # comprimido a ~1-2 GB de objetos Python y mata la instancia de Cloud Run
    # por OOM. El modo streaming sólo materializa la fila en curso. Aquí sólo
    # se usa iter_rows(values_only=True), que funciona igual en ese modo.
    # El finally es obligatorio: en read_only openpyxl deja abiertos los
    # handles del ZIP si no se cierra el workbook.
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return ParsedInput([], 0)

        headers = [str(cell or "").strip() for cell in header_row]

        rows: list[dict[str, str]] = []
        descartadas = 0
        for ws_row in rows_iter:
            row = dict(zip(headers, ws_row))
            uuid = str(row.get("UUID") or "").strip()
            # Se descarta la fila en vez de reventar el lote: un UUID malformado
            # en la fila 300 de un Excel no debe abortar las otras 499. El
            # rechazo duro vive en _require_uuid, en el único punto que arma la
            # URL de Diverza.
            if not _is_uuid(uuid):
                if uuid:
                    descartadas += 1
                continue
            rows.append(
                {
                    "uuid": uuid,
                    "rfc_emisor": str(row.get("RFC emisor") or "").strip().upper(),
                    "rfc_receptor": str(row.get("RFC receptor") or "").strip(),
                    "total_cfdi": str(row.get("TotalCFDI") or ""),
                    "motive": str(row.get("Motive") or "01"),
                }
            )
        return ParsedInput(rows, descartadas)
    finally:
        wb.close()


def _build_result_excel(
    rows_input: list[dict[str, str]], results: list[dict[str, Any] | None]
) -> bytes:
    def _sanitize_xlsx(val: str) -> str:
        if not val or not isinstance(val, str):
            return str(val) if val else ""
        if val.startswith(("=", "+", "-", "@")):
            return "'" + val
        return val

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados SAT"
    ws.append(
        [
            "UUID",
            "RFC emisor",
            "RFC receptor",
            "Motive",
            "estado",
            "es_cancelable",
            "estatus_cancelacion",
            "fecha_consulta",
            "error",
        ]
    )

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row, result in zip(rows_input, results):
        r: dict[str, Any] = result or {}
        ws.append(
            [
                _sanitize_xlsx(row["uuid"]),
                _sanitize_xlsx(row["rfc_emisor"]),
                _sanitize_xlsx(row["rfc_receptor"]),
                _sanitize_xlsx(row["motive"]),
                _sanitize_xlsx(r.get("estado", "")),
                _sanitize_xlsx(r.get("es_cancelable", "")),
                _sanitize_xlsx(r.get("estatus_cancelacion", "")),
                now,
                _sanitize_xlsx(r.get("error", "") or ""),
            ]
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post("/enquiry", response_model=EnquiryResult)
async def single_sat_enquiry(
    body: EnquiryRequest,
    tenant_id: str = Depends(verify_user_identity),
    _rate=rate_limit(20),
) -> EnquiryResult:
    cred = get_cred(body.rfc_emisor.upper(), tenant_id)
    if not cred:
        raise HTTPException(
            status_code=404,
            detail=f"RFC emisor no configurado: {body.rfc_emisor}",
        )

    payload = _build_payload(
        body.uuid, body.rfc_emisor, body.rfc_receptor, body.total_cfdi, body.motive, cred
    )

    async with httpx.AsyncClient() as client:
        try:
            text = await _call_diverza(client, body.uuid, payload)
        except httpx.HTTPError as exc:
            report(exc, contexto="consulta_sat")
            raise HTTPException(status_code=502, detail="Error al consultar el SAT") from exc

    result = _parse_diverza_response(text)
    return EnquiryResult(uuid=body.uuid, **result)


_UPLOAD_READ_CHUNK_BYTES = 64 * 1024


async def _read_xlsx_upload(file: UploadFile) -> bytes:
    """Lee el XLSX por chunks y aplica el límite antes de parsearlo."""
    chunks: list[bytes] = []
    received = 0
    while chunk := await file.read(_UPLOAD_READ_CHUNK_BYTES):
        received += len(chunk)
        if received > SAT_XLSX_MAX_BYTES:
            raise HTTPException(status_code=413, detail="El archivo excede el límite de 10 MB")
        chunks.append(chunk)
    return b"".join(chunks)


@router.post("/enquiry/batch")
async def batch_sat_enquiry(
    file: UploadFile = File(...),
    tenant_id: str = Depends(verify_user_identity),
    _rate=rate_limit(5),
) -> StreamingResponse:
    content = await _read_xlsx_upload(file)

    try:
        rows, descartadas = _parse_excel_input(content)
    except Exception as exc:
        report(exc, contexto="leer_excel")
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo de Excel") from exc

    if not rows:
        # Si TODAS las filas traían UUID pero ninguna era válida, decirlo: el
        # mensaje genérico haría pensar que el archivo venía vacío.
        if descartadas:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Ninguna de las {descartadas} filas tiene un UUID válido. "
                    "El UUID de un CFDI tiene la forma "
                    "xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx."
                ),
            )
        raise HTTPException(status_code=400, detail="El archivo no contiene filas con UUID")

    total = len(rows)
    job_id = str(uuid4())
    results: list[dict[str, Any] | None] = [None] * total

    async def event_stream():
        async with httpx.AsyncClient(
            limits=httpx.Limits(max_connections=20, max_keepalive_connections=20)
        ) as client:
            tasks = [
                asyncio.create_task(
                    _enquiry_indexed(
                        client,
                        idx,
                        row["uuid"],
                        row["rfc_emisor"],
                        row["rfc_receptor"],
                        row["total_cfdi"],
                        row["motive"],
                        tenant_id,
                    )
                )
                for idx, row in enumerate(rows)
            ]

            processed = 0
            for coro in asyncio.as_completed(tasks):
                idx, result = await coro
                results[idx] = result
                processed += 1
                yield f"data: {json.dumps({'type': 'progress', 'processed': processed, 'total': total})}\n\n"

        excel_bytes = _build_result_excel(rows, results)

        # Token de descarga: 32 bytes aleatorios, NO el job_id. El job_id sigue
        # viajando por el SSE (lo necesita el progreso) pero ya no abre nada.
        download_token = secrets.token_urlsafe(32)
        try:
            pipe = _redis_client.pipeline()
            pipe.setex(_result_key(job_id), _RESULT_TTL_SECONDS, excel_bytes)
            pipe.setex(_token_key(download_token), _RESULT_TTL_SECONDS, job_id.encode())
            pipe.execute()
        except Exception as exc:
            # Redis es la ÚNICA copia del Excel, así que un fallo aquí sí es
            # terminal para la descarga -- pero el usuario ya vio sus
            # resultados en pantalla. Se le dice que no puede bajarlo, en vez
            # de darle un token que va a dar 404.
            report(exc, contexto="guardar_resultado_lote")
            yield (
                "data: "
                + json.dumps(
                    {
                        "type": "error",
                        "message": "Los resultados se consultaron pero no se pudieron "
                        "preparar para descarga. Vuelve a intentarlo.",
                    }
                )
                + "\n\n"
            )
            return
        # `descartadas`: filas del Excel que traían un UUID malformado y no se
        # consultaron. Sin este dato el usuario sube 500 filas, recibe 480
        # resultados y no tiene forma de saber qué pasó con las otras 20.
        yield (
            "data: "
            + json.dumps(
                {
                    "type": "done",
                    "download_token": download_token,
                    "total": total,
                    "descartadas": descartadas,
                }
            )
            + "\n\n"
        )

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/enquiry/batch/result")
def get_batch_result(token: str) -> Response:
    """Descarga el Excel de un lote. Requiere el token del evento SSE `done`.

    La ruta ya no lleva el job_id: ese identificador viaja en claro por el SSE
    y no puede seguir siendo la llave de la descarga (#3).
    """
    try:
        job_id_raw = _redis_client.get(_token_key(token))
        if not job_id_raw:
            raise HTTPException(
                status_code=404, detail="Resultado no encontrado o ya descargado"
            )
        job_id = job_id_raw.decode()
        # getdel: leer y borrar en una sola operación. Con dos llamadas, dos
        # peticiones simultáneas con el mismo token bajarían las dos.
        excel_bytes = _redis_client.getdel(_result_key(job_id))
        _redis_client.delete(_token_key(token))
    except HTTPException:
        raise
    except Exception as exc:
        report(exc, contexto="descargar_resultado_lote")
        raise HTTPException(
            status_code=503, detail="No se pudo recuperar el resultado"
        ) from exc

    if not excel_bytes:
        raise HTTPException(status_code=404, detail="Resultado expirado o ya descargado")

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="consultas_sat.xlsx"'},
    )
